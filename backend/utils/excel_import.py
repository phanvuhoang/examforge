import io
import re
import logging
from typing import BinaryIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_excel(file: BinaryIO) -> list[dict]:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        while len(cells) < 5:
            cells.append(None)
        rows.append(cells)

    wb.close()

    questions = []
    i = 0

    while i < len(rows):
        col_a, col_b, col_c, col_d, col_e = rows[i]
        col_a = str(col_a).strip() if col_a else ""
        col_b_str = str(col_b).strip() if col_b else ""
        col_c_str = str(col_c).strip() if col_c else ""
        col_d_str = str(col_d).strip() if col_d else ""
        col_e_str = str(col_e).strip() if col_e else ""

        if not col_a:
            i += 1
            continue

        if col_a.upper() == "POOL":
            pool_count = int(col_b_str) if col_b_str.isdigit() else 1
            pool_questions = []
            i += 1
            while i < len(rows):
                check_a = str(rows[i][0]).strip() if rows[i][0] else ""
                if check_a.upper() == "END":
                    i += 1
                    break
                sub_questions = _parse_question_block(rows, i)
                if sub_questions:
                    q, new_i = sub_questions
                    pool_questions.append(q)
                    i = new_i
                else:
                    i += 1

            questions.append({
                "type": "TEXT",
                "body_html": f"<p>POOL: select {pool_count} from {len(pool_questions)} questions</p>",
                "body_plain": f"POOL: select {pool_count} from {len(pool_questions)} questions",
                "pool_count": pool_count,
                "pool_questions": pool_questions,
            })
            continue

        result = _parse_question_block(rows, i)
        if result:
            q, new_i = result
            questions.append(q)
            i = new_i
        else:
            i += 1

    return questions


def _parse_question_block(rows: list, start: int) -> tuple[dict, int] | None:
    if start >= len(rows):
        return None

    col_a, col_b, col_c, col_d, col_e = rows[start]
    question_text = str(col_a).strip() if col_a else ""
    col_b_str = str(col_b).strip() if col_b else ""
    col_c_str = str(col_c).strip() if col_c else ""
    col_d_str = str(col_d).strip() if col_d else ""
    col_e_str = str(col_e).strip() if col_e else ""

    if not question_text or question_text.startswith("*") or question_text.upper() in ("POOL", "END"):
        return None

    points = _try_float(col_b_str)
    explanation = col_d_str if col_d_str else None

    if col_c_str.lower() == "short" and (points is None or points == 0):
        return ({
            "type": "SA",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": 0,
            "correct_answer_json": {"accepted": [], "graded": False},
            "explanation_html": f"<p>{explanation}</p>" if explanation else "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, start + 1)

    if col_c_str.lower() == "long" and (points is None or points == 0):
        return ({
            "type": "ESSAY",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": 0,
            "correct_answer_json": {"rubric": ""},
            "explanation_html": f"<p>{explanation}</p>" if explanation else "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, start + 1)

    if col_c_str.lower() == "long" and points is not None and points > 0:
        return ({
            "type": "ESSAY",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": points,
            "correct_answer_json": {"rubric": ""},
            "explanation_html": f"<p>{explanation}</p>" if explanation else "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, start + 1)

    if col_c_str.lower() == "short" and points is not None and points > 0:
        return ({
            "type": "SA",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": points,
            "correct_answer_json": {"accepted": [], "graded": True},
            "explanation_html": f"<p>{explanation}</p>" if explanation else "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, start + 1)

    answer_rows = []
    j = start + 1
    while j < len(rows):
        next_a = str(rows[j][0]).strip() if rows[j][0] else ""
        next_b = str(rows[j][1]).strip() if rows[j][1] else ""

        if next_a.startswith("*") or next_b.startswith("~"):
            answer_rows.append(rows[j])
            j += 1
        elif next_a and not next_a.startswith("*") and next_a.upper() not in ("POOL", "END"):
            next_b_val = _try_float(str(rows[j][1]).strip() if rows[j][1] else "")
            if not next_a.startswith("*") and next_b_val is None and not str(rows[j][1] or "").strip().startswith("~"):
                if answer_rows:
                    break
                else:
                    break
            else:
                answer_rows.append(rows[j])
                j += 1
        else:
            break

    if not answer_rows and not col_b_str and not col_c_str:
        return ({
            "type": "TEXT",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": 0,
            "correct_answer_json": None,
            "explanation_html": "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, j)

    has_tilde = any(str(r[1] or "").strip().startswith("~") for r in answer_rows)
    if has_tilde:
        return _parse_matching(question_text, points or 1.0, explanation, answer_rows, col_c_str, j)

    star_rows = [r for r in answer_rows if str(r[0] or "").strip().startswith("*")]
    non_star_rows = [r for r in answer_rows if not str(r[0] or "").strip().startswith("*")]

    if star_rows and not non_star_rows:
        accepted = []
        for r in star_rows:
            val = str(r[1] or "").strip() if r[1] else str(r[0] or "").strip().lstrip("*").strip()
            if val:
                accepted.append(val)
        if len(star_rows) >= 2 and all(not r[1] for r in answer_rows):
            accepted = [str(r[0] or "").strip().lstrip("*").strip() for r in star_rows]

        return ({
            "type": "FITB",
            "body_html": f"<p>{question_text}</p>",
            "body_plain": question_text,
            "points_default": points or 1.0,
            "correct_answer_json": {"accepted": accepted},
            "explanation_html": f"<p>{explanation}</p>" if explanation else "",
            "difficulty": "medium",
            "shuffle_options": False,
            "shuffle_right_col": False,
            "options": [],
        }, j)

    options = []
    correct_count = 0
    shuffle = col_c_str.lower() != "norightshuffle"

    for idx, r in enumerate(answer_rows):
        a_val = str(r[0] or "").strip()
        b_val = str(r[1] or "").strip()
        is_correct = a_val.startswith("*")

        if is_correct:
            answer_text = b_val if b_val else a_val.lstrip("*").strip()
            correct_count += 1
            partial = _try_float(b_val.lstrip("~")) if b_val.startswith("~") else 0
        else:
            answer_text = a_val if a_val else b_val
            partial = 0

        options.append({
            "body_html": f"<p>{answer_text}</p>",
            "is_correct": is_correct,
            "display_order": idx,
            "partial_credit_pct": partial or 0,
        })

    if correct_count > 1:
        q_type = "MR"
        correct_indices = [i for i, o in enumerate(options) if o["is_correct"]]
        correct_answer = {"option_indices": correct_indices}
    elif correct_count == 1:
        if len(options) == 2 and any(
            o["body_html"].lower().replace("<p>", "").replace("</p>", "").strip() in ("true", "false", "đúng", "sai")
            for o in options
        ):
            q_type = "TF"
            correct_opt = next(o for o in options if o["is_correct"])
            val_text = correct_opt["body_html"].lower().replace("<p>", "").replace("</p>", "").strip()
            correct_answer = {"value": val_text in ("true", "đúng")}
        else:
            q_type = "MC"
            correct_idx = next(i for i, o in enumerate(options) if o["is_correct"])
            correct_answer = {"option_index": correct_idx}
    else:
        q_type = "MC"
        correct_answer = {}

    return ({
        "type": q_type,
        "body_html": f"<p>{question_text}</p>",
        "body_plain": question_text,
        "points_default": points or 1.0,
        "correct_answer_json": correct_answer,
        "explanation_html": f"<p>{explanation}</p>" if explanation else "",
        "difficulty": "medium",
        "shuffle_options": shuffle,
        "shuffle_right_col": True,
        "options": options,
    }, j)


def _parse_matching(
    question_text: str,
    points: float,
    explanation: str | None,
    answer_rows: list,
    col_c_str: str,
    end_idx: int,
) -> tuple[dict, int]:
    pairs = []
    options = []
    for idx, r in enumerate(answer_rows):
        b_val = str(r[1] or "").strip()
        c_val = str(r[2] or "").strip()

        left_text = b_val.lstrip("~").strip() if b_val.startswith("~") else b_val
        right_text = c_val

        pairs.append({"left": left_text, "right": right_text})
        options.append({
            "body_html": f"<p>{left_text} → {right_text}</p>",
            "is_correct": True,
            "display_order": idx,
            "partial_credit_pct": 0,
        })

    return ({
        "type": "MATCH",
        "body_html": f"<p>{question_text}</p>",
        "body_plain": question_text,
        "points_default": points,
        "correct_answer_json": {"pairs": pairs},
        "explanation_html": f"<p>{explanation}</p>" if explanation else "",
        "difficulty": "medium",
        "shuffle_options": True,
        "shuffle_right_col": col_c_str.lower() != "norightshuffle",
        "options": options,
    }, end_idx)


def _try_float(val: str) -> float | None:
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
